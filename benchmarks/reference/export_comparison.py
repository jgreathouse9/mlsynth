#!/usr/bin/env python3
"""Export side-by-side mlsynth-vs-reference comparison tables.

A cross-validation case defines ``comparison() -> {"rows": [...],
"mlsynth_call": {...}}``: the rows pair the library's number against the captured
reference value quantity by quantity, and ``mlsynth_call`` records the estimator
and config the mlsynth numbers came from. This writes, per case, a committed,
metadata-stamped artifact::

    benchmarks/reference/<case>/comparison.csv   # metadata header + quantity, mlsynth, reference, abs_diff

and a combined workbook -- the headline a reviewer opens -- with a summary sheet
plus one detail sheet per case::

    benchmarks/reference/comparisons.xlsx

Each artifact carries provenance: when it was written, the mlsynth version and
the exact call (estimator + config), and the reference implementation and its
version.

Usage::

    python benchmarks/reference/export_comparison.py synth_prop99
    python benchmarks/reference/export_comparison.py --all
"""
from __future__ import annotations

import argparse
import csv
import datetime as _dt
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
REF_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

import mlsynth                                  # noqa: E402
from benchmarks import registry                # noqa: E402

_FIELDS = ["quantity", "mlsynth", "reference", "abs_diff"]
# Fixed workbook doc-properties timestamp -> byte-deterministic .xlsx output.
_EPOCH = _dt.datetime(2020, 1, 1, tzinfo=_dt.timezone.utc)


def _cases_with_comparison() -> list:
    out = []
    for name in registry.CASES:
        try:
            mod = registry.load(name)
        except Exception:                        # pragma: no cover - import-time issues
            continue
        if hasattr(mod, "comparison"):
            out.append(name)
    return out


def _missing_cases() -> list:
    """Comparison cases that do not yet have a committed comparison.csv.

    The daily action generates only these -- re-running an existing (often
    stochastic) case would churn Monte-Carlo jitter into the committed workbook.
    """
    return [n for n in _cases_with_comparison()
            if not (REF_DIR / n / "comparison.csv").exists()]


def _read_comparison_csv(path: Path) -> dict:
    """Parse a committed comparison.csv back into ``{"rows", "meta"}``.

    The metadata header is ``# key: value`` comment lines; the body is the
    ``quantity, mlsynth, reference, abs_diff`` table. Numeric columns are coerced
    to float so the assembled workbook matches a live export.
    """
    meta: dict = {}
    body: list = []
    for line in Path(path).read_text().splitlines():
        if line.startswith("# "):
            key, _, val = line[2:].partition(": ")
            meta[key] = val
        elif line.strip():
            body.append(line)
    rows = []
    for r in csv.DictReader(body):
        rows.append({"quantity": r["quantity"],
                     "mlsynth": float(r["mlsynth"]),
                     "reference": float(r["reference"]),
                     "abs_diff": float(r["abs_diff"])})
    return {"rows": rows, "meta": meta}


def assemble_workbook_from_csv(ref_dir: Path = REF_DIR) -> dict:
    """Rebuild comparisons.xlsx from every committed comparison.csv under ref_dir.

    Assembling from the committed CSVs (rather than re-running each case) lets one
    job produce the complete workbook even though the cases live in incompatible
    reference environments, and keeps entries stable day to day.
    """
    per_case = {}
    for csv_path in sorted(ref_dir.glob("*/comparison.csv")):
        per_case[csv_path.parent.name] = _read_comparison_csv(csv_path)
    _write_workbook(per_case, out_dir=ref_dir)
    return per_case


def _metadata(name: str, result: dict) -> dict:
    """Provenance for one case's comparison: when, which mlsynth call, which
    reference (and version). The reference descriptor comes from the case's
    ``comparison()`` (for live, vendored references) or from the captured bundle's
    manifest/provenance when present."""
    call = result.get("mlsynth_call", {})
    meta = {
        "case": name,
        "generated_at": _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "mlsynth_version": getattr(mlsynth, "__version__", "unknown"),
        "estimator": call.get("estimator", ""),
        "config": json.dumps(call.get("config", {}), sort_keys=True, default=str),
    }
    ref = result.get("reference", {})
    if ref:                                       # case-supplied (live reference)
        meta["reference_impl"] = ref.get("impl", "")
        meta["reference_version"] = ref.get("version", "")
        return meta
    bundle = REF_DIR / name                       # captured-bundle reference
    if (bundle / "manifest.json").exists():
        meta["reference_impl"] = json.loads((bundle / "manifest.json").read_text()).get("reference_impl", "")
    if (bundle / "provenance.json").exists():
        prov = json.loads((bundle / "provenance.json").read_text())
        ver = prov.get("r_version") or prov.get("python_version") or ""
        pkgs = prov.get("packages", {})
        key = next((p for p in ("Synth", "augsynth", "scpi", "causaltensor") if p in pkgs), None)
        if key:
            ver = f"{ver}; {key} {pkgs[key]}".strip("; ")
        meta["reference_version"] = ver
        meta["reference_generated_at"] = prov.get("generated_at", "")
    return meta


def export_case(name: str) -> dict:
    """Run a case's comparison(), write its comparison.csv, return rows + metadata."""
    result = registry.load(name).comparison()
    rows = result["rows"]
    for r in rows:
        r["abs_diff"] = round(abs(float(r["mlsynth"]) - float(r["reference"])), 6)
    meta = _metadata(name, result)
    bundle = REF_DIR / name
    bundle.mkdir(exist_ok=True)
    with open(bundle / "comparison.csv", "w", newline="") as fh:
        for k, v in meta.items():                # metadata as leading comment lines
            fh.write(f"# {k}: {v}\n")
        w = csv.DictWriter(fh, fieldnames=_FIELDS)
        w.writeheader()
        w.writerows(rows)
    print(f"[ok] {name}: wrote comparison.csv ({len(rows)} rows, max |Δ|="
          f"{max(r['abs_diff'] for r in rows):.4g})")
    return {"rows": rows, "meta": meta}


def _write_workbook(per_case: dict, out_dir: Path = REF_DIR) -> None:
    """Combined workbook: a summary sheet plus a metadata-stamped sheet per case."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:                          # pragma: no cover - openpyxl present here
        print("[note] openpyxl not installed; skipping comparisons.xlsx (CSVs written)")
        return
    wb = Workbook()
    # Pin the workbook's own doc-properties timestamps so identical content yields
    # byte-identical output -- otherwise openpyxl stamps the current time and the
    # daily "commit only if changed" gate would churn the binary every run.
    wb.properties.created = _EPOCH
    wb.properties.modified = _EPOCH
    summ = wb.active
    summ.title = "summary"
    summ.append(["case", "quantities", "max_abs_diff", "generated_at",
                 "mlsynth_version", "reference_impl"])
    for c in summ[1]:
        c.font = Font(bold=True)
    for name, payload in per_case.items():
        rows, meta = payload["rows"], payload["meta"]
        summ.append([name, len(rows), max(r["abs_diff"] for r in rows),
                     meta["generated_at"], meta["mlsynth_version"],
                     meta.get("reference_impl", "")])
        ws = wb.create_sheet(name[:31])
        for k, v in meta.items():                # metadata block atop the sheet
            ws.append([k, v])
            ws.cell(ws.max_row, 1).font = Font(bold=True)
        ws.append([])
        ws.append(_FIELDS)
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
        for r in rows:
            ws.append([r[f] for f in _FIELDS])
    out = out_dir / "comparisons.xlsx"
    wb.save(out)
    _normalize_xlsx_bytes(out)
    print(f"[ok] wrote {out} (summary + {len(per_case)} case sheet(s))")


def _normalize_xlsx_bytes(path: Path) -> None:
    """Rewrite an .xlsx so identical content is byte-identical.

    openpyxl stamps every zip member with the wall-clock time (2-second DOS
    granularity), so two saves of the same workbook differ. Repackage the members
    in sorted order under a fixed timestamp -- with the pinned doc-properties this
    makes the file a pure function of its content, so the daily commit-if-changed
    gate stays quiet unless an entry actually changed.
    """
    import re
    import zipfile

    fixed = (1980, 1, 1, 0, 0, 0)                 # earliest DOS zip epoch
    stamp = "2020-01-01T00:00:00Z"
    with zipfile.ZipFile(path) as src:
        members = sorted(src.namelist())
        data = {m: src.read(m) for m in members}
    # openpyxl rewrites dcterms:modified to now() at save regardless of the set
    # property, so pin created/modified in core.xml to a constant here.
    core = data.get("docProps/core.xml")
    if core is not None:
        text = core.decode("utf-8")
        for tag in ("created", "modified"):
            text = re.sub(rf"(<dcterms:{tag}[^>]*>)[^<]*(</dcterms:{tag}>)",
                          rf"\g<1>{stamp}\g<2>", text)
        data["docProps/core.xml"] = text.encode("utf-8")
    tmp = path.with_name(path.name + ".tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as dst:
        for m in members:
            zi = zipfile.ZipInfo(m, date_time=fixed)
            zi.compress_type = zipfile.ZIP_DEFLATED
            dst.writestr(zi, data[m])
    tmp.replace(path)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("case", nargs="?")
    ap.add_argument("--all", action="store_true",
                    help="(re)run every case that defines comparison()")
    ap.add_argument("--missing", action="store_true",
                    help="run only cases that have no committed comparison.csv yet")
    ap.add_argument("--assemble", action="store_true",
                    help="rebuild comparisons.xlsx from committed CSVs; no live runs")
    args = ap.parse_args()

    if args.assemble:                             # pure assembly, no reference runs
        per_case = assemble_workbook_from_csv()
        print(f"[ok] assembled comparisons.xlsx from {len(per_case)} committed CSV(s)")
        return 0

    if args.missing:
        cases = _missing_cases()
    elif args.all:
        cases = _cases_with_comparison()
    else:
        cases = [args.case] if args.case else []
    if not cases:
        if args.missing:                          # nothing to do is success, not an error
            print("[ok] no missing comparison entries")
            return 0
        ap.error("give a case name, --all, --missing, or --assemble "
                 "(a case must define comparison())")
    from benchmarks.compare import BenchmarkSkipped
    per_case = {}
    for name in cases:
        try:
            per_case[name] = export_case(name)
        except BenchmarkSkipped as exc:
            print(f"[skip] {name}: {exc}")
        except Exception as exc:                  # reference toolchain absent / draft error
            print(f"[skip] {name}: {type(exc).__name__}: {str(exc)[:160]}")
    # Rebuild the workbook from ALL committed CSVs (the ones just written plus the
    # pre-existing ones), so a partial run never drops another case's entry.
    assemble_workbook_from_csv()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
